import os
import re
import json
import openpyxl

def compile_staff():
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'LOLOS STAFF IB2K26.xlsx')
    excel_path = os.path.abspath(excel_path)
    print(f"Reading Excel file from: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found!")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 1. Parse Contact Persons
    cps = {}
    if 'CONTACT PERSON' in wb.sheetnames:
        cp_sheet = wb['CONTACT PERSON']
        for row in cp_sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            subdiv = str(row[0]).strip().lower()
            cp_str = row[1] or ''
            # Match pattern wa.me/([0-9]+) \((.+)\) or just extract number and name
            match = re.search(r'wa\.me/([0-9]+)\s*\((.+)\)', cp_str)
            if match:
                phone = match.group(1).strip()
                name = match.group(2).strip()
            else:
                phone = ''
                name = cp_str.strip()
            cps[subdiv] = {'phone': phone, 'name': name}
    
    print("Parsed Contact Persons:", json.dumps(cps, indent=2))

    # Map sheet names to contact person subdivisions
    sheet_to_cp_map = {
        'DAMEN': 'data management',
        'COMPE': 'competition',
        'CEREMONY': 'ceremony',
        'PUBLIC RELATION': 'public relation',
        'SNL': 'snl',
        'LTE': 'lte',
        'MEDICAL': 'medical',
        'CONSUMPTION': 'consumption',
        'SPONSORSHIP': 'sponsorship',
        'TICKETING': 'ticketing',
        'FUNDRAISE': 'fundraising',
        'BRANDING ': 'branding',
        'CND': 'creative',
        'MEDPRO': 'medpro',
        'UIUX': 'ui/ux',
        'FRONT END': 'front-end',
        'BACK END': 'back-end'
    }

    # Pretty subdivision names for display on the frontend
    pretty_subdivision_names = {
        'data management': 'Data Management',
        'competition': 'Competition',
        'ceremony': 'Ceremony',
        'public relation': 'Public Relation',
        'snl': 'SnL',
        'lte': 'LTE',
        'medical': 'Medical',
        'consumption': 'Consumption',
        'sponsorship': 'Sponsorship',
        'ticketing': 'Ticketing',
        'fundraising': 'Fundraising',
        'branding': 'Branding',
        'creative': 'CnD',
        'medpro': 'MedPro',
        'ui/ux': 'UI/UX',
        'front-end': 'Front-End',
        'back-end': 'Back-End'
    }

    results = {}
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'CONTACT PERSON':
            continue
            
        ws = wb[sheet_name]
        cp_key = sheet_to_cp_map.get(sheet_name)
        cp_info = cps.get(cp_key, {'phone': '', 'name': ''})
        pretty_subdiv = pretty_subdivision_names.get(cp_key, sheet_name)
        
        rows = list(ws.iter_rows(values_only=True))
        data_started = False
        nama_idx = -1
        nrp_idx = -1
        
        for row in rows:
            if not row:
                continue
            
            row_str = [str(x).strip().lower() if x is not None else '' for x in row]
            if 'nrp' in ''.join(row_str) or 'nama' in ''.join(row_str):
                data_started = True
                for idx, cell in enumerate(row_str):
                    if 'nama' in cell:
                        nama_idx = idx
                    elif 'nrp' in cell:
                        nrp_idx = idx
                continue
            
            if data_started:
                if nama_idx != -1 and nrp_idx != -1 and len(row) > max(nama_idx, nrp_idx):
                    nama_val = row[nama_idx]
                    nrp_val = row[nrp_idx]
                    if nama_val is None or nrp_val is None:
                        continue
                    
                    nama = str(nama_val).strip()
                    nrp = str(nrp_val).strip()
                    if nrp.endswith('.0'):
                        nrp = nrp[:-2]
                    
                    # Remove non-digits from NRP to normalize it
                    nrp = re.sub(r'\D', '', nrp)
                    if not nrp or len(nrp) < 5:
                        continue
                    
                    # Add candidate to results map by NRP
                    results[nrp] = {
                        'nama': nama,
                        'nrp': nrp,
                        'subdivisi': pretty_subdiv,
                        'cp_name': cp_info['name'],
                        'cp_phone': cp_info['phone']
                    }

    print(f"Total passed staff compiled: {len(results)}")
    
    # Save to constants directory
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'constants')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'lolos_staff.json')
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
        
    print(f"Database successfully saved to: {output_path}")

if __name__ == '__main__':
    compile_staff()
