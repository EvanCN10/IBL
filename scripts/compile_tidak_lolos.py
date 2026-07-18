import os
import re
import json
import openpyxl

def compile_tidak_lolos():
    """
    Reads 'Staff Interview IBL2K26.xlsx' to extract ALL interviewed staff (Nama + NRP).
    Then cross-references with 'lolos_staff.json' to produce 'tidak_lolos_staff.json'
    containing only staff who did NOT pass.

    Handles the case where a single Excel tab contains multiple divisions
    laid out side-by-side (horizontally), each with their own Nama/NRP columns.
    """

    base_dir = os.path.join(os.path.dirname(__file__), '..')
    excel_path = os.path.abspath(os.path.join(base_dir, 'Staff Interview IBL2K26.xlsx'))
    lolos_path = os.path.abspath(os.path.join(base_dir, 'frontend', 'constants', 'lolos_staff.json'))

    print(f"Reading interview Excel from: {excel_path}")
    print(f"Reading lolos staff from:     {lolos_path}")

    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found!")
        return

    if not os.path.exists(lolos_path):
        print(f"Error: {lolos_path} not found!")
        return

    # Load the set of NRPs that already passed
    with open(lolos_path, 'r', encoding='utf-8') as f:
        lolos_data = json.load(f)
    lolos_nrps = set(lolos_data.keys())
    print(f"Loaded {len(lolos_nrps)} passed staff NRPs from lolos_staff.json")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_staff = {}  # NRP -> { "nama": ... }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Processing sheet: '{sheet_name}' ---")

        # Read all rows as a 2D list for easier manipulation
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            print(f"  Sheet '{sheet_name}' is empty, skipping.")
            continue

        # Find the header row: the row that contains cells with 'nama' and 'nrp'
        header_row_idx = -1
        for idx, row in enumerate(all_rows):
            row_lower = [str(cell).strip().lower() if cell is not None else '' for cell in row]
            has_nama = any('nama' in cell for cell in row_lower)
            has_nrp = any('nrp' in cell for cell in row_lower)
            if has_nama and has_nrp:
                header_row_idx = idx
                break

        if header_row_idx == -1:
            print(f"  Could not find header row with 'Nama' and 'NRP' in sheet '{sheet_name}', skipping.")
            continue

        header_row = all_rows[header_row_idx]
        header_lower = [str(cell).strip().lower() if cell is not None else '' for cell in header_row]

        # Find ALL (nama_col, nrp_col) pairs in this header row
        # This handles multiple divisions laid out side-by-side
        nama_cols = []
        nrp_cols = []
        for col_idx, cell_val in enumerate(header_lower):
            if 'nama' in cell_val:
                nama_cols.append(col_idx)
            elif 'nrp' in cell_val:
                nrp_cols.append(col_idx)

        # Pair each nama column with the nearest nrp column to its right
        pairs = []
        used_nrp_cols = set()
        for nama_col in nama_cols:
            best_nrp_col = None
            best_distance = float('inf')
            for nrp_col in nrp_cols:
                if nrp_col in used_nrp_cols:
                    continue
                distance = abs(nrp_col - nama_col)
                if distance < best_distance:
                    best_distance = distance
                    best_nrp_col = nrp_col
            if best_nrp_col is not None:
                pairs.append((nama_col, best_nrp_col))
                used_nrp_cols.add(best_nrp_col)

        print(f"  Found {len(pairs)} division group(s) in header row {header_row_idx + 1}")

        # Extract data from rows below the header
        data_rows = all_rows[header_row_idx + 1:]
        count_in_sheet = 0
        for row in data_rows:
            for nama_col, nrp_col in pairs:
                if len(row) <= max(nama_col, nrp_col):
                    continue

                nama_val = row[nama_col]
                nrp_val = row[nrp_col]

                if nama_val is None or nrp_val is None:
                    continue

                nama = str(nama_val).strip()
                nrp = str(nrp_val).strip()

                # Remove trailing .0 from Excel number formatting
                if nrp.endswith('.0'):
                    nrp = nrp[:-2]

                # Normalize NRP: keep only digits
                nrp = re.sub(r'\D', '', nrp)

                # Skip invalid/too-short NRPs
                if not nrp or len(nrp) < 5:
                    continue

                # Skip empty names
                if not nama:
                    continue

                all_staff[nrp] = {'nama': nama}
                count_in_sheet += 1

        print(f"  Extracted {count_in_sheet} staff entries from sheet '{sheet_name}'")

    print(f"\nTotal interviewed staff extracted: {len(all_staff)}")

    # Filter out staff who already passed (are in lolos_staff.json)
    tidak_lolos = {}
    for nrp, data in all_staff.items():
        if nrp not in lolos_nrps:
            tidak_lolos[nrp] = data

    print(f"Staff who did NOT pass: {len(tidak_lolos)}")
    print(f"(Total interviewed: {len(all_staff)} - Passed: {len(all_staff) - len(tidak_lolos)} = Not passed: {len(tidak_lolos)})")

    # Save output
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'constants')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'tidak_lolos_staff.json')

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(tidak_lolos, f, indent=2, ensure_ascii=False)

    print(f"\nDatabase saved to: {output_path}")


if __name__ == '__main__':
    compile_tidak_lolos()
